"""
표와 텍스트 감지 및 추출 관련 유틸리티 함수
"""

import os
import cv2
import numpy as np
import logging
import tempfile
import csv
import openpyxl
import io
import re
from PIL import Image
from typing import List, Dict, Any, Tuple, Optional, Union
import tabula
import easyocr
import pandas as pd

# 로깅 설정
logger = logging.getLogger(__name__)

# JPype 로깅 레벨 설정
logging.getLogger('jpype').setLevel(logging.ERROR)

# 전역 OCR 리더 저장 변수
_ocr_reader = None

# 셀 크롭 마진 상수 정의 (셀 이미지 자를 때 여백)
CELL_CROP_MARGIN = 2

# 표 감지 관련 상수
class TableConfig:
    # 이미지 처리 설정
    MAX_WIDTH = 1800
    MIN_WIDTH = 800
    
    # 표 감지 설정
    MIN_LINE_LENGTH = 100
    MAX_LINE_GAP = 10
    CELL_MARGIN = 5
    
    # 그리드 감지 설정
    ROW_THRESHOLD = 20
    COL_THRESHOLD = 30
    
    # OCR 설정
    OCR_LANGS = ['ko', 'en']
    
    # 디버그 모드 설정
    DEBUG = True
    DEBUG_OUTPUT_DIR = 'debug_images'
    
    # PDF 관련 설정
    PDF_LATTICE = True
    PDF_GUESS = False
    PDF_TABLE_AREA = None
    PDF_PAGES = 'all'
    
    # 출력 형식 설정
    DEFAULT_OUTPUT_FORMAT = 'csv'
    
    # 텍스트 블록 감지 설정
    MIN_TEXT_HEIGHT = 20
    MIN_TEXT_WIDTH = 30
    TEXT_MARGIN = 10


# OCR 리더 객체 초기화 (첫 호출시만 생성)
def get_ocr_reader():
    """EasyOCR 리더 객체를 가져오는 함수"""
    global _ocr_reader
    if _ocr_reader is None:
        logger.info("EasyOCR 초기화 중...")
        _ocr_reader = easyocr.Reader(TableConfig.OCR_LANGS, gpu=False)
        logger.info("EasyOCR 초기화 완료")
    return _ocr_reader


# 디버그 모드인 경우 디렉토리 생성
if TableConfig.DEBUG and not os.path.exists(TableConfig.DEBUG_OUTPUT_DIR):
    os.makedirs(TableConfig.DEBUG_OUTPUT_DIR)


def resize_image(image: np.ndarray, max_width: int = TableConfig.MAX_WIDTH) -> np.ndarray:
    """이미지 크기 조정"""
    height, width = image.shape[:2]
    
    # 이미지가 너무 작으면 확대
    if width < TableConfig.MIN_WIDTH:
        scale = TableConfig.MIN_WIDTH / width
        new_width = int(width * scale)
        new_height = int(height * scale)
        return cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_CUBIC)
    
    # 이미지가 너무 크면 축소
    if width > max_width:
        scale = max_width / width
        new_width = int(width * scale)
        new_height = int(height * scale)
        return cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
    
    return image


def preprocess_image(image: np.ndarray) -> np.ndarray:
    """이미지 전처리 - 개선된 버전"""
    # 그레이스케일 변환
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
    
    # 노이즈 제거
    denoised = cv2.fastNlMeansDenoising(gray)
    
    # 적응형 이진화
    binary = cv2.adaptiveThreshold(
        denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 19, 3
    )
    
    # 모폴로지 연산으로 선 강화
    kernel = np.ones((2, 2), np.uint8)
    binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    
    return binary


def detect_grid_structure(cells):
    """셀들의 격자 구조 감지"""
    if not cells:
        return [], []
    
    # Y 좌표로 행 그룹화
    y_coords = [cell[1] for cell in cells]
    y_coords.extend([cell[3] for cell in cells])
    y_coords = sorted(list(set(y_coords)))
    
    # X 좌표로 열 그룹화
    x_coords = [cell[0] for cell in cells]
    x_coords.extend([cell[2] for cell in cells])
    x_coords = sorted(list(set(x_coords)))
    
    # 비슷한 좌표 병합
    rows = [y_coords[0]]
    for y in y_coords[1:]:
        if y - rows[-1] > TableConfig.ROW_THRESHOLD:
            rows.append(y)
            
    cols = [x_coords[0]]
    for x in x_coords[1:]:
        if x - cols[-1] > TableConfig.COL_THRESHOLD:
            cols.append(x)
    
    return rows, cols


def assign_cells_to_grid(cells, rows, cols):
    """셀들을 격자에 할당"""
    grid = [['' for _ in range(len(cols)-1)] for _ in range(len(rows)-1)]
    
    for cell in cells:
        x1, y1, x2, y2 = cell
        
        # 셀이 속한 행과 열 찾기
        row_idx = None
        col_idx = None
        
        for i in range(len(rows)-1):
            if rows[i] <= (y1 + y2)/2 <= rows[i+1]:
                row_idx = i
                break
                
        for j in range(len(cols)-1):
            if cols[j] <= (x1 + x2)/2 <= cols[j+1]:
                col_idx = j
                break
        
        if row_idx is not None and col_idx is not None:
            grid[row_idx][col_idx] = cell
    
    return grid


def extract_text_from_cell(image: np.ndarray, cell: Tuple[int, int, int, int]) -> str:
    """셀 이미지에서 텍스트 추출 - OCR 결과 중 첫 번째 항목만 사용"""
    x1, y1, x2, y2 = cell
    
    # 여백 추가 (감소된 마진 사용)
    margin = CELL_CROP_MARGIN  # 마진은 작은 값 유지
    y1 = max(0, y1 - margin)
    y2 = min(image.shape[0], y2 + margin)
    x1 = max(0, x1 - margin)
    x2 = min(image.shape[1], x2 + margin)
    
    # 영역 유효성 확인
    if y1 >= y2 or x1 >= x2:
        logger.warning(f"잘못된 셀 좌표 또는 크기 (0): {cell}")
        return ""
    
    cell_img = image[y1:y2, x1:x2]
    
    # 이미지 유효성 추가 확인
    if cell_img is None or cell_img.size == 0:
        logger.warning(f"빈 셀 이미지 또는 로드 실패: {cell}")
        return ""
    
    # 이미지 전처리
    if len(cell_img.shape) == 3:
        gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY)
    else:
        gray = cell_img
    
    # 이미지 크기 조정
    try:
        scale = 1.5  # 스케일을 조금 줄임
        # 리사이징 전에 그레이 이미지 크기 확인
        if gray.size > 0:
            cell_img_resized = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        else:
            logger.warning(f"리사이징 불가: 셀 이미지 크기 0. 좌표: {cell}")
            return ""
    except cv2.error as resize_error:
        logger.error(f"OpenCV 리사이징 오류: {resize_error}, 셀 좌표: {cell}")
        # 리사이징 실패 시 원본 그레이 이미지 사용
        cell_img_resized = gray
    
    # OCR 실행
    try:
        reader = get_ocr_reader()
        results = reader.readtext(cell_img_resized, detail=0, paragraph=False)
        
        # --- 핵심 수정: 첫 번째 결과만 사용 ---
        if results:
            # 여러 텍스트 조각이 감지되어도 첫 번째 것만 해당 셀의 내용으로 간주
            first_result = results[0].strip()
            # 간단한 후처리 (연속 공백 제거)
            cleaned_text = re.sub(r'\s+', ' ', first_result)
            logger.debug(f"셀 {cell} OCR 결과 (첫 번째): '{cleaned_text}' (원본 결과: {results})")  # 디버깅 로그
            return cleaned_text
        else:
            logger.debug(f"셀 {cell} OCR 결과 없음")
            return ""  # OCR 결과가 없으면 빈 문자열 반환
        # --- 수정 완료 ---
        
    except Exception as ocr_error:
        logger.exception(f"OCR 실행 중 오류 발생: {ocr_error}, 셀 좌표: {cell}")
        return "[OCR 오류]"  # OCR 오류 시 표시


def detect_table_cells(binary: np.ndarray) -> List[Tuple[int, int, int, int]]:
    """표 셀 감지 - 단순화된 버전"""
    # 수직선과 수평선 강화
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 25))
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 1))
    
    # 수직선 감지
    vertical_lines = cv2.erode(binary, vertical_kernel, iterations=1)
    vertical_lines = cv2.dilate(vertical_lines, vertical_kernel, iterations=1)
    
    # 수평선 감지
    horizontal_lines = cv2.erode(binary, horizontal_kernel, iterations=1)
    horizontal_lines = cv2.dilate(horizontal_lines, horizontal_kernel, iterations=1)
    
    # 모든 선 결합
    table_grid = cv2.addWeighted(vertical_lines, 1, horizontal_lines, 1, 0)
    
    # 교차점 강화
    table_grid = cv2.dilate(table_grid, np.ones((3,3), np.uint8), iterations=1)
    
    # 셀 윤곽선 감지
    contours, _ = cv2.findContours(table_grid, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    # 셀 좌표 추출
    cells = []
    min_cell_area = 100  # 최소 셀 크기
    
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if w * h > min_cell_area:
            cells.append((x, y, x+w, y+h))
    
    # 셀 정렬 (위에서 아래로, 왼쪽에서 오른쪽으로)
    cells.sort(key=lambda x: (x[1], x[0]))
    
    return cells


def organize_cells_into_table(cells: List[Tuple[int, int, int, int]], image_shape: Tuple[int, int]) -> List[List[Tuple[int, int, int, int]]]:
    """셀들을 표 형식으로 구성"""
    if not cells:
        return []
    
    # Y 좌표로 행 그룹화
    row_tolerance = 10
    current_row = [cells[0]]
    rows = []
    
    for cell in cells[1:]:
        if abs(cell[1] - current_row[0][1]) <= row_tolerance:
            current_row.append(cell)
        else:
            # 현재 행의 셀들을 X 좌표로 정렬
            current_row.sort(key=lambda x: x[0])
            rows.append(current_row)
            current_row = [cell]
    
    if current_row:
        current_row.sort(key=lambda x: x[0])
        rows.append(current_row)
    
    return rows


def extract_table_from_image(image_path: str) -> Dict[str, Any]:
    """이미지에서 표 추출 - 개선된 버전"""
    try:
        # 이미지 로드
        image = cv2.imread(image_path)
        if image is None:
            return {'success': False, 'error': '이미지를 불러올 수 없습니다'}
        
        # 그레이스케일 변환
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # 적응형 이진화
        binary = cv2.adaptiveThreshold(
            gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV, 19, 3
        )
        
        if TableConfig.DEBUG:
            cv2.imwrite(os.path.join(TableConfig.DEBUG_OUTPUT_DIR, 'binary.png'), binary)
        
        # 셀 감지
        cells = detect_table_cells(binary)
        if not cells:
            return {'success': False, 'error': '표를 감지할 수 없습니다'}
        
        # 셀들을 표 형식으로 구성
        table_rows = organize_cells_into_table(cells, image.shape[:2])
        
        # OCR 수행
        table_data = []
        for row in table_rows:
            row_data = []
            for cell_coords in row:  # 변수명을 cell_coords로 변경하여 명확화
                # extract_text_from_cell 호출 시 전체 이미지와 셀 좌표를 전달
                text = extract_text_from_cell(image, cell_coords)  # 수정된 부분
                row_data.append(text.strip())
            table_data.append(row_data)
        
        return {
            'success': True,
            'table_data': table_data,
            'cells_count': len(cells)
        }
        
    except Exception as e:
        # 상세한 오류 로깅
        logger.exception(f"extract_table_from_image 함수에서 오류 발생: {str(e)}")
        return {'success': False, 'error': str(e)}


def save_table_to_csv(table_data: List[List[str]], output_path: str) -> bool:
    """표 데이터를 CSV 파일로 저장"""
    try:
        # CSV 파일에 UTF-8-SIG BOM을 포함하여 저장 (Excel 호환성)
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # 첫 번째 행은 헤더로 간주
            if table_data:
                writer.writerow(table_data[0])
                # 중복 행 제거 로직 추가
                seen_rows = set()
                unique_rows = []
                
                for row in table_data[1:]:
                    # 튜플로 변환하여 해시 가능하게 만듦
                    row_tuple = tuple(str(cell) for cell in row)
                    if row_tuple not in seen_rows and not all(cell == '' for cell in row):
                        seen_rows.add(row_tuple)
                        unique_rows.append(row)
                
                writer.writerows(unique_rows)
                logger.info(f"CSV 저장 완료: {output_path} (중복 제거 후 {len(unique_rows)}행)")
            else:
                logger.warning(f"저장할 테이블 데이터가 없음: {output_path}")
        return True
    except Exception as e:
        logger.error(f"CSV 저장 중 오류 발생: {str(e)}")
        return False


def save_table_to_excel(table_data: List[List[str]], output_path: str) -> bool:
    """표 데이터를 Excel 파일로 저장"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 헤더 스타일 설정
        header_style = openpyxl.styles.NamedStyle(name='header')
        header_style.font = openpyxl.styles.Font(bold=True)
        header_style.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # 테두리 스타일
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )
        
        # 중복 행 제거 로직 추가
        unique_rows = []
        seen_rows = set()
        
        # 헤더 추가
        if table_data:
            unique_rows.append(table_data[0])
            
            # 데이터 행 처리 (중복 제거)
            for row in table_data[1:]:
                row_tuple = tuple(str(cell) for cell in row)
                if row_tuple not in seen_rows and not all(cell == '' for cell in row):
                    seen_rows.add(row_tuple)
                    unique_rows.append(row)
        
        # 데이터 추가 및 스타일 적용
        for i, row in enumerate(unique_rows):
            for j, cell_value in enumerate(row):
                # 숫자형 데이터 변환 시도
                cell = ws.cell(row=i+1, column=j+1)
                
                # 숫자 형식 자동 감지 (통화, 백분율, 일반 숫자)
                try:
                    # 쉼표, 통화 기호, 공백 제거
                    clean_value = str(cell_value).replace(',', '').replace('₩', '').replace('$', '').strip()
                    
                    # 퍼센트 처리
                    if clean_value.endswith('%'):
                        numeric_value = float(clean_value.rstrip('%')) / 100
                        cell.value = numeric_value
                        cell.number_format = '0.00%'
                    # 일반 숫자 처리
                    elif clean_value and re.match(r'^-?\d+(\.\d+)?$', clean_value):
                        numeric_value = float(clean_value)
                        cell.value = numeric_value
                        # 소수점이 있는지 확인
                        if '.' in clean_value:
                            # 소수점 이하 자리수 유지
                            decimal_places = len(clean_value.split('.')[1])
                            cell.number_format = f'0.{"0" * decimal_places}'
                        else:
                            cell.number_format = '0'
                    else:
                        cell.value = cell_value
                except (ValueError, TypeError):
                    # 변환 실패 시 문자열로 저장
                    cell.value = cell_value
                
                # 테두리 스타일 적용
                cell.border = thin_border
                
                # 헤더 스타일 적용
                if i == 0:
                    cell.style = header_style
        
        # 열 너비 자동 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 30)  # 최대 폭 제한
        
        wb.save(output_path)
        logger.info(f"Excel 저장 완료: {output_path} (중복 제거 후 {len(unique_rows)-1}행)")
        return True
    except Exception as e:
        logger.error(f"Excel 저장 중 오류 발생: {str(e)}")
        return False


def generate_csv_in_memory(table_data: List[List[str]]) -> bytes:
    """표 데이터를 메모리에서 CSV 문자열(utf-8-sig 인코딩 바이트)로 생성"""
    output = io.StringIO()
    try:
        writer = csv.writer(output)
        writer.writerows(table_data)
        # utf-8-sig로 인코딩하여 바이트 반환 (BOM 포함)
        return output.getvalue().encode('utf-8-sig')
    except Exception as e:
        logger.error(f"CSV 메모리 생성 중 오류: {e}")
        # 오류 발생 시 빈 바이트 데이터 반환
        return b""
    finally:
        output.close()  # StringIO 객체 닫기


def generate_excel_in_memory(table_data: List[List[str]]) -> bytes:
    """표 데이터를 메모리에서 Excel 바이트 데이터로 생성"""
    output = io.BytesIO()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        for i, row in enumerate(table_data):
            for j, cell_value in enumerate(row):
                # 숫자형 데이터 변환 시도 (Excel에서 숫자로 인식되도록)
                try:
                    # 쉼표 제거 후 숫자 변환 시도
                    numeric_value = float(cell_value.replace(',', ''))
                    ws.cell(row=i+1, column=j+1, value=numeric_value)
                except (ValueError, TypeError):
                    # 변환 실패 시 문자열로 저장
                    ws.cell(row=i+1, column=j+1, value=cell_value)
        
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        logger.error(f"Excel 메모리 생성 중 오류: {e}")
        return b""
    finally:
        output.close()  # BytesIO 객체 닫기


def extract_table_data_as_content(
    file_path: str,
    original_filename: str,  # 원본 파일명 인자 추가
    output_format: str = TableConfig.DEFAULT_OUTPUT_FORMAT
) -> Dict[str, Any]:
    """파일에서 표를 추출하고, 결과를 메모리 내 컨텐츠로 반환 (헤더 후처리 제거)"""
    file_ext = os.path.splitext(file_path)[1].lower()
    is_pdf = file_ext == '.pdf'

    logger.info(f"표 추출 시작: 원본 파일={original_filename}, 임시 파일={os.path.basename(file_path)}, PDF={is_pdf}")

    if is_pdf:
        result = extract_table_from_pdf(file_path)
    else:
        result = extract_table_from_image(file_path)

    # Java 오류 감지 및 사용자 친화적인 오류 메시지
    if not result or not result.get('success'):
        error_msg = result.get('error', '알 수 없는 오류') if result else '결과 없음'
        logger.error(f"표 추출 실패: {error_msg}, 파일={original_filename}")
        
        # Java 관련 오류 감지
        if result and isinstance(error_msg, str) and ("JVM mismatch" in error_msg or "UnsupportedClassVersionError" in error_msg):
            # 대체 안내 테이블 생성
            fallback_table = [
                ["표 추출을 위한 Java 설정 안내"],
                [""],
                ["현재 오류: Java 버전/아키텍처 불일치"],
                [""],
                ["해결 방법:"],
                ["1. Eclipse Temurin JDK 8 이상(64비트) 설치"],
                ["   - https://adoptium.net/temurin/releases/?version=8"],
                ["2. 설치 후 서버 재시작"],
                ["3. JAVA_HOME 환경 변수가 올바르게 설정되었는지 확인"],
                [""],
                ["또는 표 이미지를 직접 캡처하여 처리하세요."],
            ]
            
            # 오류에 대한 추가 정보 제공
            if "64 bit" in error_msg and "32 bit" in error_msg:
                fallback_table.append([""])
                fallback_table.append(["자세한 오류 정보: 64비트 Python에 32비트 Java가 설치되어 있습니다."])
                fallback_table.append(["64비트 Java로 업데이트하세요."])
            elif "UnsupportedClassVersionError" in error_msg:
                fallback_table.append([""])
                fallback_table.append(["자세한 오류 정보: Java 버전이 너무 낮습니다."])
                fallback_table.append(["Java 8 이상이 필요합니다."])
            
            logger.info("Java 오류에 대한 대체 안내 테이블 생성")
            
            # 파일 생성
            base_name = os.path.splitext(original_filename)[0]
            
            try:
                if output_format.lower() == 'excel':
                    file_content = generate_excel_in_memory(fallback_table)
                    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    filename = f"{base_name}_table_info.xlsx"
                else:  # 기본값 CSV
                    file_content = generate_csv_in_memory(fallback_table)
                    mime_type = 'text/csv; charset=utf-8-sig'
                    filename = f"{base_name}_table_info.csv"
                
                return {
                    'success': True,
                    'file_content': file_content,
                    'filename': filename,
                    'mime_type': mime_type,
                    'is_fallback': True,
                    'table_data': fallback_table,
                    'error': error_msg
                }
            except Exception as fallback_error:
                logger.error(f"대체 안내 테이블 생성 실패: {fallback_error}")
        
        return {
            'success': False,
            'error': f"표 추출 실패: {error_msg}"
        }

    table_data = result.get('table_data')
    # 테이블 데이터가 리스트 형태인지, 비어있지 않은지 확인
    if not isinstance(table_data, list) or not table_data:
        logger.warning(f"표 추출은 성공했으나 유효한 테이블 데이터가 없음: 파일={original_filename}")
        return {
            'success': False,
            'error': '추출된 표 데이터가 없습니다.'
        }

    # ------------------------------------------
    # --- 헤더 후처리 로직 (복원 및 개선) ---
    # ------------------------------------------
    # 표 데이터가 충분히 있는지 확인
    if len(table_data) > 1:
        # 헤더 후보 분석
        header_candidates = []
        
        # 첫 번째 행을 헤더로 간주
        header_row = table_data[0]
        data_rows = table_data[1:]
        
        # 첫 행이 실제 헤더인지 확인 (모든 데이터 행과 다른 형태인지)
        is_header_likely = False
        
        # 숫자만 있는 헤더 감지 (숫자 데이터일 가능성)
        all_numeric = True
        for header in header_row:
            if not re.match(r'^[\d\.\s\-\,]+$', str(header).strip()):
                all_numeric = False
                break
        
        # 첫 행이 다른 행들과 얼마나 다른지 체크
        if len(data_rows) > 0:
            # 첫 행과 데이터 행의 형식이 다른지 확인
            header_format = [bool(re.search(r'[a-zA-Z가-힣]', str(h))) for h in header_row]
            data_format = [
                [bool(re.search(r'[a-zA-Z가-힣]', str(cell))) for cell in row]
                for row in data_rows[:3]  # 처음 3개 행만 확인
            ]
            
            # 형식 차이 계산
            format_diff = sum(h != sum(d[i] for d in data_format)/len(data_format) > 0.5 
                             for i, h in enumerate(header_format)) / len(header_format)
            
            is_header_likely = format_diff > 0.5 or not all_numeric
        
        # 첫 행이 헤더일 가능성이 낮으면 column_1, column_2 형식으로 생성
        if not is_header_likely and all_numeric:
            logger.info("첫 행이 헤더로 적합하지 않아 보입니다 (숫자로만 구성). 자동 헤더 생성.")
            generated_header = [f"Column_{i+1}" for i in range(len(header_row))]
            # 원래 첫 행을 데이터에 포함
            processed_data = [header_row] + data_rows
            header_row = generated_header
        else:
            # 헤더 후처리: 빈 셀, 중복 헤더, 특수문자 처리
            processed_header = []
            column_names = set()
            
            for i, header in enumerate(header_row):
                # 빈 헤더 처리
                if not header or header.strip() == '':
                    header = f'Column_{i+1}'
                
                # 특수문자 제거 및 공백 정리
                header = re.sub(r'[^\w\s\(\)\.\-]', '', str(header)).strip()
                
                # 중복 헤더 처리
                if header in column_names:
                    header = f'{header}_{i+1}'
                
                column_names.add(header)
                processed_header.append(header)
            
            header_row = processed_header
            processed_data = data_rows
        
        # 데이터 행 후처리
        clean_data = []
        for row in processed_data:
            # 빈 행 건너뛰기 (모든 셀이 비어있는 경우)
            if all(not cell or str(cell).strip() == '' for cell in row):
                continue
            
            # 행 길이 맞추기 (헤더보다 짧은 경우 빈 셀 추가)
            if len(row) < len(header_row):
                row = row + [''] * (len(header_row) - len(row))
            # 헤더보다 긴 경우 잘라내기
            elif len(row) > len(header_row):
                row = row[:len(header_row)]
            
            # 각 셀 정리: 앞뒤 공백 제거, None 처리, 줄바꿈 제거
            processed_row = []
            for cell in row:
                if cell is None:
                    processed_row.append('')
                else:
                    # 줄바꿈과 연속된 공백을 단일 공백으로 변환
                    clean_cell = re.sub(r'\s+', ' ', str(cell).strip())
                    processed_row.append(clean_cell)
            
            clean_data.append(processed_row)
        
        # 중복 행 제거 (연속된 동일 행 제거)
        deduplicated_data = []
        prev_row = None
        for row in clean_data:
            row_tuple = tuple(row)
            if row_tuple != prev_row:
                deduplicated_data.append(row)
                prev_row = row_tuple
        
        # 처리된 헤더와 데이터를 결합
        table_data = [header_row] + deduplicated_data
        logger.info(f"헤더 후처리 완료: {header_row}")
    else:
        logger.warning("데이터가 충분하지 않아 헤더 후처리 건너뜀")

    try:
        file_content: Union[str, bytes]
        mime_type: str
        filename: str
        # 원본 파일명에서 확장자를 제외한 부분을 사용
        base_name = os.path.splitext(original_filename)[0]

        if output_format.lower() == 'excel':
            file_content = generate_excel_in_memory(table_data)
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"{base_name}_table.xlsx"
            # 생성된 Excel 데이터 크기 로깅
            logger.debug(f"생성된 Excel 데이터 크기: {len(file_content)} 바이트")
        else:  # 기본값 CSV
            file_content = generate_csv_in_memory(table_data)
            # 생성된 CSV 데이터 크기 로깅
            logger.debug(f"생성된 CSV 데이터 크기: {len(file_content)} 바이트")
            if not file_content:  # 생성 실패 시 오류 처리
                raise ValueError("CSV 데이터 생성 실패 (빈 결과)")
            # CSV는 utf-8-sig 를 명시
            mime_type = 'text/csv; charset=utf-8-sig'
            filename = f"{base_name}_table.csv"

        # 최종 반환 전 데이터 로깅 (디버깅용)
        logger.info(f"파일 컨텐츠 생성 완료: 파일명={filename}, 타입={mime_type}, 원본 파일={original_filename}")

        return {
            'success': True,
            'file_content': file_content,
            'filename': filename,
            'mime_type': mime_type,
            'cells_count': result.get('cells_count', 0) or result.get('tables_count', 0)
        }

    except Exception as e:
        logger.exception(f"파일 컨텐츠 생성 중 오류 발생: {str(e)}, 파일={original_filename}")
        return {
            'success': False,
            'error': f"표 데이터를 {output_format} 형식으로 변환하는 중 오류 발생: {str(e)}"
        }


def extract_table_from_pdf(pdf_path: str) -> Dict[str, Any]:
    """PDF에서 표 직접 추출"""
    try:
        # Tabula 설정
        lattice_mode = TableConfig.PDF_LATTICE
        guess = TableConfig.PDF_GUESS
        area = TableConfig.PDF_TABLE_AREA
        pages = TableConfig.PDF_PAGES
        
        logger.info(f"PDF에서 표 추출 시작: {pdf_path}")
        
        # 첫 번째 10페이지만 스캔하기 위한 설정
        total_pages = 1
        try:
            import PyPDF2
            with open(pdf_path, 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                total_pages = len(pdf_reader.pages)
                
                # 페이지 수가 너무 많으면 처음 10페이지로 제한
                if isinstance(pages, str) and pages == 'all' and total_pages > 10:
                    logger.info(f"PDF가 너무 큽니다 ({total_pages} 페이지). 처음 10 페이지만 스캔합니다.")
                    pages = '1-10'
                    
                # 텍스트 컨텐츠 확인 (표를 포함하고 있는지)
                has_text = False
                for i in range(min(3, total_pages)):  # 처음 3 페이지만 확인
                    page = pdf_reader.pages[i]
                    text = page.extract_text()
                    if text and len(text.strip()) > 50:  # 최소 텍스트 길이
                        has_text = True
                        break
                
                if not has_text:
                    logger.warning("PDF에 텍스트가 충분하지 않습니다. 스캔된 문서일 수 있으므로 이미지 처리를 시도합니다.")
                    raise ValueError("텍스트가 충분하지 않은 PDF")
                
        except Exception as pdf_error:
            logger.warning(f"PDF 메타데이터 분석 실패: {pdf_error}. 전체 페이지를 스캔합니다.")
        
        # 캐싱 - 이전에 처리된 파일이 있는지 확인
        import os
        cache_key = f"{pdf_path}_{lattice_mode}_{pages}"
        cache_file = os.path.join(os.path.dirname(pdf_path), ".table_cache", os.path.basename(pdf_path) + ".json")
        
        if os.path.exists(cache_file):
            try:
                import json
                with open(cache_file, 'r', encoding='utf-8') as f:
                    cache_data = json.load(f)
                    if cache_data.get('cache_key') == cache_key:
                        logger.info(f"캐시된 결과를 사용합니다: {cache_file}")
                        return {
                            'success': True,
                            'table_data': cache_data.get('table_data', []),
                            'tables_count': cache_data.get('tables_count', 0),
                            'from_cache': True
                        }
            except Exception as cache_error:
                logger.warning(f"캐시 로드 실패: {cache_error}")

        # Java 버전 확인 함수
        def get_java_version():
            try:
                import subprocess
                # capture_output과 stdout/stderr를 함께 사용하면 충돌 발생
                # 대신 pipe를 사용하고 communicate()로 출력 캡처
                process = subprocess.Popen(
                    ["java", "-version"], 
                    stdout=subprocess.PIPE, 
                    stderr=subprocess.PIPE,
                    universal_newlines=True
                )
                stdout, stderr = process.communicate()
                
                # java -version은 stderr에 출력함
                version_output = stderr or stdout
                
                # Java 버전 문자열 파싱 (예: "java version "1.8.0_301"")
                import re
                version_match = re.search(r'version "([^"]+)"', version_output)
                if version_match:
                    version_str = version_match.group(1)
                    # 메이저 버전 추출
                    if version_str.startswith("1."):  # 1.8 형식 (Java 8 이하)
                        major_version = int(version_str.split(".")[1])
                    else:  # 9, 10, 11 등 형식 (Java 9 이상)
                        major_version = int(version_str.split(".")[0])
                    
                    logger.info(f"감지된 Java 버전: {version_str} (메이저 버전: {major_version})")
                    
                    # 아키텍처 확인 (Client VM = 32bit, Server VM = 64bit)
                    vm_match = re.search(r'(Client|Server) VM', version_output)
                    is_client_vm = vm_match and vm_match.group(1) == 'Client'
                    
                    if is_client_vm:
                        logger.warning("32비트 Java VM이 감지되었습니다. 64비트 Java가 필요합니다.")
                        # 강제로 Java 버전을 낮춰서 대체 방법 사용
                        return 0
                    
                    return major_version
                return 0  # 버전을 확인할 수 없음
            except Exception as e:
                logger.warning(f"Java 버전 확인 실패: {e}")
                return 0  # 버전을 확인할 수 없음

        # Java 버전에 따른 옵션 설정
        java_version = get_java_version()
        
        # 기본 Java 옵션
        java_options = [
            "-Xmx2048m",
            "-Xms512m"
        ]
        
        # Java 버전별 GC 옵션 설정
        if java_version >= 9:
            logger.info("Java 9+ 감지, 최신 JVM 옵션 적용")
            java_options.append("-XX:+UseStringDeduplication")
        elif java_version >= 8:
            logger.info("Java 8 감지, Java 8 최적화 옵션 적용")
            java_options.append("-XX:+UseG1GC")
            java_options.append("-XX:+UseStringDeduplication")
        elif java_version >= 7:
            logger.info("Java 7 감지, Java 7 호환 옵션 적용")
            java_options.append("-XX:+UseG1GC")
        else:
            logger.info("Java 6 또는 버전 확인 불가, 기본 JVM 옵션만 적용")
        
        # 모든 버전에서 공통으로 사용하는 PDF 관련 설정
        java_options.extend([
            "-Dsun.java2d.cmm=sun.java2d.cmm.kcms.KcmsServiceProvider",  # 색상 관리
            "-Dorg.apache.pdfbox.rendering.UsePureJavaCMYKConversion=true",  # PDF 렌더링
            "-Dorg.apache.pdfbox.rendering.UsePureJava=true",  # 순수 Java 렌더링
            "-Dorg.apache.pdfbox.rendering.ignoreMissingUnicode=true"  # 유니코드 매핑 무시
        ])
        
        # ===== 방법 1: 직접 subprocess를 사용하는 방식 =====
        try:
            # 직접 pandas DataFrame 생성
            import pandas as pd
            import subprocess
            import tempfile
            import json
            import os
            
            logger.info("직접 subprocess를 사용하여 표 추출을 시도합니다.")
            logger.info(f"사용할 Java 옵션: {' '.join(java_options)}")
            
            # 임시 출력 파일 생성
            with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as temp:
                output_path = temp.name
            
            try:
                # Tabula JAR 파일 찾기
                import pkg_resources
                tabula_jar = None
                for resource in pkg_resources.working_set:
                    if resource.key == 'tabula-py':
                        tabula_jar = os.path.join(resource.location, 'tabula', 'tabula-1.0.5-jar-with-dependencies.jar')
                        break
                
                if not tabula_jar or not os.path.exists(tabula_jar):
                    # 패키지 설치 위치 기본 경로에서 검색
                    import importlib
                    tabula_module = importlib.import_module('tabula')
                    tabula_path = os.path.dirname(tabula_module.__file__)
                    tabula_jar = os.path.join(tabula_path, 'tabula-1.0.5-jar-with-dependencies.jar')
                
                if not os.path.exists(tabula_jar):
                    logger.error(f"Tabula JAR 파일을 찾을 수 없습니다: {tabula_jar}")
                    raise FileNotFoundError(f"Tabula JAR 파일을 찾을 수 없습니다")
                
                # java 명령어 확인
                java_cmd = "java"
                
                if os.name == 'nt':
                    # 가능한 64비트 Java 경로 목록 확장
                    java64_paths = [
                        r"C:\Program Files\Java\jdk-18.0.2.1\bin\java.exe"
                    ]
                    
                    import glob
                    for pattern in java64_paths:
                        matches = glob.glob(pattern)
                        if matches:
                            # 가장 최신 버전 선택 (버전 번호로 정렬)
                            matches.sort(reverse=True)  # 높은 버전 번호가 앞에 오도록
                            java_cmd = matches[0]
                            logger.info(f"64비트 Java 경로 사용: {java_cmd}")
                            break
                    
                    # 환경 변수에서 JAVA_HOME 확인
                    java_home = os.environ.get('JAVA_HOME')
                    if not matches and java_home:
                        java_path = os.path.join(java_home, 'bin', 'java.exe')
                        if os.path.exists(java_path):
                            java_cmd = java_path
                            logger.info(f"JAVA_HOME에서 Java 경로 사용: {java_cmd}")

                # 명령 구성
                cmd = [java_cmd]
                cmd.extend(java_options)
                cmd.extend(["-jar", tabula_jar])
                
                # 페이지 옵션
                if pages != 'all':
                    cmd.extend(["--pages", str(pages)])
                
                # 기타 옵션
                if lattice_mode:
                    cmd.append("--lattice")
                else:
                    cmd.append("--stream")
                
                # 출력 형식
                cmd.extend(["--format", "JSON", "--outfile", output_path])
                
                # 입력 파일
                cmd.append(pdf_path)
                
                # 명령 실행 (타임아웃 설정)
                logger.info(f"명령 실행: {' '.join(cmd)}")
                
                process = subprocess.Popen(
                    cmd, 
                    stdout=subprocess.PIPE, 
                    stderr=subprocess.PIPE,
                    universal_newlines=True
                )
                
                # 120초 타임아웃으로 프로세스 대기
                import time
                start_time = time.time()
                timeout = 120  # 초
                
                while process.poll() is None:
                    if time.time() - start_time > timeout:
                        process.kill()
                        raise TimeoutError("표 추출 시간 초과 (120초)")
                    time.sleep(0.1)
                
                stdout, stderr = process.communicate()
                
                if process.returncode != 0:
                    logger.error(f"Tabula 명령 실패: {stderr}")
                    
                    # 스트림 모드로 재시도
                    if lattice_mode:
                        logger.info("격자 모드 실패, 스트림 모드로 재시도합니다")
                        
                        # 격자 모드 비활성화
                        cmd = [opt for opt in cmd if opt != "--lattice"]
                        cmd.append("--stream")
                        cmd.append("--guess")  # 스트림 모드에서는 guess=True가 더 나은 결과를 줄 수 있음
                        
                        logger.info(f"재시도 명령 실행: {' '.join(cmd)}")
                        
                        process = subprocess.Popen(
                            cmd, 
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE,
                            universal_newlines=True
                        )
                        
                        # 120초 타임아웃으로 프로세스 대기
                        start_time = time.time()
                        
                        while process.poll() is None:
                            if time.time() - start_time > timeout:
                                process.kill()
                                raise TimeoutError("스트림 모드 표 추출 시간 초과 (120초)")
                            time.sleep(0.1)
                        
                        stdout, stderr = process.communicate()
                        
                        if process.returncode != 0:
                            logger.error(f"스트림 모드에서도 Tabula 명령 실패: {stderr}")
                            raise Exception(f"표 추출 명령 실패: {stderr}")
                
                # JSON 파일 로드
                if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
                    logger.error("Tabula가 출력 파일을 생성하지 않았습니다")
                    raise FileNotFoundError("Tabula 출력 파일을 찾을 수 없습니다")
                
                with open(output_path, 'r', encoding='utf-8') as f:
                    json_tables = json.load(f)
                
                # 결과가 비어있는 경우
                if not json_tables:
                    logger.warning("PDF에서 표를 찾을 수 없습니다")
                    raise ValueError("PDF에서 표를 감지할 수 없습니다")
                
                # DataFrame 목록으로 변환
                dfs = []
                for table in json_tables:
                    if 'data' in table:
                        # JSON 배열을 DataFrame으로 변환
                        data = table['data']
                        if data:
                            df = pd.DataFrame(data)
                            # 각 셀에서 'text' 키 추출
                            df = df.applymap(lambda cell: cell.get('text', '') if isinstance(cell, dict) else '')
                            dfs.append(df)
                
                # 결과가 비어있는 경우
                if not dfs:
                    logger.warning("PDF에서 유효한 표를 찾을 수 없습니다")
                    raise ValueError("PDF에서 유효한 표를 감지할 수 없습니다")
                
            except Exception as subprocess_error:
                logger.error(f"Subprocess 처리 중 오류: {subprocess_error}")
                raise
            finally:
                # 임시 파일 삭제
                if os.path.exists(output_path):
                    try:
                        os.unlink(output_path)
                    except Exception as e:
                        logger.warning(f"임시 파일 삭제 실패: {e}")
        
        except Exception as direct_subprocess_error:
            logger.error(f"직접 subprocess 호출 실패: {direct_subprocess_error}")
            logger.info("대체 방법으로 tabula-py API 호출을 시도합니다 (TABULA_USE_SUBPROCESS=true)")
            
            # ===== 방법 2: tabula-py를 통한 호출 =====
            try:
                # tabula-py를 통한 호출 (환경 변수로 subprocess 모드 강제)
                import os
                os.environ["TABULA_USE_SUBPROCESS"] = "true"
                
                # tabula 모듈을 명시적으로 다시 로드하여 환경 변수 변경 적용
                import sys
                import importlib
                if 'tabula' in sys.modules:
                    importlib.reload(sys.modules['tabula'])
                
                # Java 환경 변수 설정 - 버전별 옵션 적용 (Java 8 미만에서도 작동하는 옵션만 사용)
                safe_java_options = [
                    "-Xmx2048m", 
                    "-Xms512m",
                    "-XX:+UseG1GC",  # Java 7 이상에서는 문제 없음
                    "-Dsun.java2d.cmm=sun.java2d.cmm.kcms.KcmsServiceProvider"
                ]
                
                # Java 8 이상일 때만 추가 옵션 포함
                if java_version >= 8:
                    safe_java_options.append("-XX:+UseStringDeduplication")
                
                os.environ["JAVA_OPTS"] = " ".join(safe_java_options)
                
                # 페이지별 처리를 통한 시도
                dfs = []
                
                # 페이지 범위 파싱
                if isinstance(pages, str):
                    if pages == 'all':
                        page_nums = list(range(1, min(total_pages + 1, 11)))  # 최대 10 페이지
                    elif '-' in pages:
                        start, end = map(int, pages.split('-'))
                        page_nums = list(range(start, min(end + 1, start + 10)))  # 최대 10 페이지
                    else:
                        page_nums = [int(p) for p in pages.split(',')][:10]  # 최대 10 페이지
                else:
                    page_nums = [int(pages)]
                
                # 각 페이지 개별 처리
                import tabula
                jvm_error = None
                
                # 두 가지 모드 시도: 격자 모드와 스트림 모드
                modes = [
                    {"lattice": True, "guess": False, "name": "격자 모드"},
                    {"lattice": False, "guess": True, "name": "스트림 모드"}
                ]
                
                all_dfs = []
                successful_mode = None
                
                for mode in modes:
                    mode_dfs = []
                    mode_error = None
                    
                    logger.info(f"{mode['name']}로 표 추출 시도")
                    
                    for page_num in page_nums:
                        try:
                            logger.info(f"페이지 {page_num} 처리 중 ({mode['name']})...")
                            page_dfs = tabula.read_pdf(
                                pdf_path,
                                pages=str(page_num),
                                area=area,
                                guess=mode["guess"],
                                lattice=mode["lattice"],
                                multiple_tables=True,
                                pandas_options={'header': None}
                            )
                            
                            if page_dfs and len(page_dfs) > 0:
                                # 빈 표 또는 너무 작은 표 필터링
                                filtered_dfs = [df for df in page_dfs 
                                               if df.shape[0] > 1 and df.shape[1] > 1]
                                
                                if filtered_dfs:
                                    mode_dfs.extend(filtered_dfs)
                                    logger.info(f"페이지 {page_num}에서 {len(filtered_dfs)}개 표 추출 성공")
                                else:
                                    logger.warning(f"페이지 {page_num}에서 유효한 표를 발견하지 못함")
                            else:
                                logger.warning(f"페이지 {page_num}에서 표를 발견하지 못함")
                                
                        except Exception as page_error:
                            # JVM 아키텍처 불일치 오류 확인
                            error_str = str(page_error)
                            logger.warning(f"페이지 {page_num} 처리 실패: {error_str}")
                            
                            if "JVM mismatch" in error_str or "UnsupportedClassVersionError" in error_str:
                                logger.error("Java 아키텍처 또는 버전 불일치 감지됨. 이미지 추출 방식으로 즉시 전환합니다.")
                                jvm_error = page_error
                                break  # 더 이상 시도하지 않고 루프 종료
                            
                            mode_error = page_error
                    
                    # JVM 오류가 발생한 경우 모든 처리 중단
                    if jvm_error:
                        break
                        
                    # 현재 모드에서 표를 발견했는지 확인
                    if mode_dfs:
                        all_dfs.extend(mode_dfs)
                        successful_mode = mode['name']
                        logger.info(f"{mode['name']}에서 총 {len(mode_dfs)}개 표 추출 성공")
                        break  # 성공했으면 다른 모드 시도하지 않음
                
                # JVM 오류가 발생했거나 모든 모드가 실패한 경우
                if jvm_error or not all_dfs:
                    if jvm_error:
                        logger.error(f"Java 오류로 인해 PDF 직접 추출 실패: {jvm_error}")
                    else:
                        logger.warning("모든 표 추출 모드가 실패했습니다.")
                        
                    # PDF를 이미지로 변환하여 재시도
                    raise ValueError("PDF 직접 추출 실패, 이미지 변환 방식으로 전환")
                
                # 성공적으로 추출된 표가 있음
                dfs = all_dfs
                logger.info(f"총 {len(dfs)}개의 표를 {successful_mode}로 추출했습니다")
                
            except Exception as tabula_error:
                error_str = str(tabula_error)
                logger.error(f"tabula-py 처리 실패: {error_str}")
                
                # JVM 관련 오류인 경우 추가 메시지 출력
                if "JVM mismatch" in error_str or "UnsupportedClassVersionError" in error_str:
                    logger.error("64비트 Java 8 이상 설치 필요. Oracle JDK 또는 Eclipse Temurin 설치 권장.")
                
                logger.info("PDF를 이미지로 변환한 후 추출을 시도합니다.")
                
                # ===== 방법 3: PDF를 이미지로 변환한 후 처리 =====
                try:
                    import tempfile
                    import pdf2image
                    import os
                    import pandas as pd
                    from pdf2image import convert_from_path
                    
                    logger.info(f"PDF를 이미지로 변환: {pdf_path}")
                    
                    # 임시 디렉토리 생성
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # PDF -> 이미지 변환 (첫 10페이지만)
                        images = convert_from_path(
                            pdf_path, 
                            dpi=300,  # 해상도 높임
                            first_page=page_nums[0] if page_nums else 1,
                            last_page=page_nums[-1] if page_nums else 10
                        )
                        
                        logger.info(f"PDF에서 {len(images)}개 이미지 생성됨")
                        
                        # 모든 이미지에서 표 추출 시도
                        all_table_data = []
                        
                        for i, img in enumerate(images):
                            img_path = os.path.join(temp_dir, f"page_{i+1}.png")
                            img.save(img_path, "PNG")
                            
                            logger.info(f"이미지 변환 후 표 추출 시도: {img_path}")
                            
                            # 이미지에서 표 추출
                            result = extract_table_from_image(img_path)
                            
                            if result.get('success') and result.get('table_data'):
                                # 추출 성공
                                all_table_data.extend(result['table_data'])
                        
                        if all_table_data:
                            # 성공: 모든 이미지에서 추출한 데이터 합치기 
                            dfs = [pd.DataFrame(all_table_data)]
                            logger.info(f"이미지 변환 방식으로 표 {len(all_table_data)}행 추출 성공")
                        else:
                            # 실패: 빈 스케줄이라도 반환
                            basic_table = [["표가 발견되지 않았습니다. 이미지 및 PDF 변환 모두 실패."]]
                            dfs = [pd.DataFrame(basic_table)]
                            logger.warning("이미지 변환 후에도 표를 감지하지 못했습니다. 기본 메시지 반환.")
                
                except Exception as img_error:
                    # 모든 대체 방식이 실패한 경우
                    logger.error(f"이미지 변환 방식도 실패: {img_error}")
                    
                    # 기본 빈 표 제공
                    logger.warning("모든 방식이 실패하여 기본 표를 제공합니다.")
                    error_table = [
                        ["PDF 표 추출 실패"],
                        [f"오류 메시지: {str(img_error)}"],
                        [""],
                        ["해결 방법:"],
                        ["1. 64비트 Java 설치 여부 확인"],
                        ["2. PDF를 이미지로 변환 후 다시 시도"],
                        ["3. 다른 PDF 뷰어에서 표를 직접 복사하여 사용"]
                    ]
                    dfs = [pd.DataFrame(error_table)]
        
        # 추출된 표 필터링 및 처리
        filtered_dfs = []
        for df in dfs:
            # 너무 작은 표 필터링 (1행 1열 같은 잘못 감지된 표)
            if df.shape[0] > 1 and df.shape[1] > 1:
                # 결측값 처리 및 문자열 변환 최적화
                filtered_dfs.append(df.fillna(''))
        
        if not filtered_dfs:
            logger.warning("유효한 표가 없습니다 (모두 필터링됨)")
            # 모든 방법이 실패한 경우 기본 안내 메시지 포함된 표 반환
            error_table = [
                ["PDF 표 추출 실패 - 유효한 표가 감지되지 않음"],
                ["가능한 해결 방법:"],
                ["1. 다른 PDF 뷰어에서 표를 직접 복사하여 사용"],
                ["2. PDF를 이미지로 변환 후 다시 시도"],
                ["3. 시스템에 64비트 Java가 설치되어 있는지 확인"]
            ]
            filtered_dfs = [pd.DataFrame(error_table)]
        
        # 추출된 모든 표 중 가장 큰 표 선택 (행 * 열)
        largest_df = max(filtered_dfs, key=lambda df: df.shape[0] * df.shape[1])
        
        # 빈 행과 중복 행 제거
        # 모든 셀이 빈 문자열인 행 제거
        largest_df = largest_df.loc[~(largest_df == '').all(axis=1)]
        
        # 중복 행 제거
        largest_df = largest_df.drop_duplicates()
        
        # 데이터 정제
        # 문자열 변환 최적화
        table_data = []
        for _, row in largest_df.iterrows():
            row_data = []
            for val in row:
                if val == '':
                    row_data.append('')
                else:
                    # 문자열 정리 (앞뒤 공백 및 특수문자 제거)
                    cleaned_val = str(val).strip()
                    # 불필요한 줄바꿈 및 공백 처리
                    cleaned_val = re.sub(r'\s+', ' ', cleaned_val)
                    row_data.append(cleaned_val)
            table_data.append(row_data)
        
        # 캐시 저장
        try:
            import json
            os.makedirs(os.path.dirname(cache_file), exist_ok=True)
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'cache_key': cache_key,
                    'table_data': table_data,
                    'tables_count': len(filtered_dfs)
                }, f, ensure_ascii=False)
            logger.info(f"캐시 저장 완료: {cache_file}")
        except Exception as cache_error:
            logger.warning(f"캐시 저장 실패: {cache_error}")
        
        return {
            'success': True,
            'table_data': table_data,
            'tables_count': len(filtered_dfs)
        }
        
    except Exception as e:
        logger.error(f"PDF 표 추출 중 오류 발생: {str(e)}")
        
        # 대체 정보 제공
        fallback_table = [
            ["PDF 표 추출 중 오류가 발생했습니다"],
            [f"오류 메시지: {str(e)}"],
            [""],
            ["해결 방법:"],
            ["1. 64비트 Java 설치 여부 확인"],
            ["2. PDF를 이미지로 변환 후 다시 시도"],
            ["3. 다른 PDF 뷰어에서 표를 직접 복사하여 사용"]
        ]
        
        return {
            'success': True,  # 사용자에게 오류 대신 도움말을 보여주기 위해 success를 True로 설정
            'table_data': fallback_table,
            'tables_count': 1,
            'is_fallback': True  # 이것이 대체 정보임을 표시
        }


def detect_text_blocks(image: np.ndarray, table_regions: List[Tuple[int, int, int, int]]) -> List[Tuple[int, int, int, int]]:
    """표 외부의 텍스트 블록 감지"""
    # 이미지 전처리
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                 cv2.THRESH_BINARY_INV, 19, 3)
    
    # 표 영역 마스킹
    mask = np.ones_like(binary) * 255
    for x1, y1, x2, y2 in table_regions:
        mask[y1:y2, x1:x2] = 0
    
    # 마스크 적용
    binary_masked = cv2.bitwise_and(binary, mask)
    
    # 텍스트 블록 감지
    contours, _ = cv2.findContours(binary_masked, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    text_blocks = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if w > TableConfig.MIN_TEXT_WIDTH and h > TableConfig.MIN_TEXT_HEIGHT:
            text_blocks.append((x, y, x+w, y+h))
    
    return text_blocks


def detect_table_regions(binary: np.ndarray) -> List[Tuple[int, int, int, int]]:
    """표 영역 감지"""
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    table_regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if w > 100 and h > 100:  # 최소 표 크기
            table_regions.append((x, y, x+w, y+h))
    
    return table_regions


def detect_content_regions(image: np.ndarray) -> Tuple[List[Tuple[int, int, int, int]], List[Tuple[int, int, int, int]]]:
    """이미지에서 표 영역과 텍스트 영역을 감지"""
    # 이미지 전처리
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
    
    # 적응형 이진화
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 19, 3
    )
    
    # 노이즈 제거
    kernel = np.ones((2, 2), np.uint8)
    binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
    
    # 윤곽선 감지
    contours, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    table_regions = []
    text_regions = []
    
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        aspect_ratio = w / float(h)
        
        # 표 영역 조건
        if area > 10000 and 0.5 < aspect_ratio < 2.0:
            table_regions.append((x, y, x+w, y+h))
        # 텍스트 영역 조건
        elif 20 < h < 100 and w > 30:
            text_regions.append((x, y, x+w, y+h))
    
    # 겹치는 영역 처리
    text_regions = [region for region in text_regions 
                   if not any(is_region_inside(region, table) for table in table_regions)]
    
    return table_regions, text_regions


def is_region_inside(region1: Tuple[int, int, int, int], region2: Tuple[int, int, int, int]) -> bool:
    """한 영역이 다른 영역 안에 있는지 확인"""
    x1, y1, x2, y2 = region1
    X1, Y1, X2, Y2 = region2
    return X1 <= x1 and Y1 <= y1 and X2 >= x2 and Y2 >= y2


def extract_content_from_image(image_path: str) -> Dict[str, Any]:
    """이미지에서 표와 텍스트 추출 - 개선된 버전"""
    try:
        # 이미지 로드
        image = cv2.imread(image_path)
        if image is None:
            return {'success': False, 'error': '이미지를 불러올 수 없습니다'}
        
        # 이미지 크기 조정
        image = resize_image(image)
        
        # 표와 텍스트 영역 감지
        table_regions, text_regions = detect_content_regions(image)
        
        # 콘텐츠 요소 저장
        content_elements = []
        
        # 텍스트 영역 처리
        for region in text_regions:
            x1, y1, x2, y2 = region
            text_img = image[y1:y2, x1:x2]
            text = extract_text_from_region(text_img)
            if text.strip():
                content_elements.append({
                    'type': 'text',
                    'y_position': y1,
                    'data': text.strip(),
                    'region': region
                })
        
        # 표 영역 처리
        for region in table_regions:
            x1, y1, x2, y2 = region
            table_img = image[y1:y2, x1:x2]
            table_result = extract_table_from_image_region(table_img)
            if table_result['success']:
                content_elements.append({
                    'type': 'table',
                    'y_position': y1,
                    'data': table_result['table_data'],
                    'region': region
                })
        
        # Y 좌표 기준으로 정렬
        content_elements.sort(key=lambda x: x['y_position'])
        
        return {
            'success': True,
            'content': content_elements
        }
        
    except Exception as e:
        logger.error(f"콘텐츠 추출 중 오류 발생: {str(e)}")
        return {'success': False, 'error': str(e)}


def extract_table_from_image_region(image: np.ndarray) -> Dict[str, Any]:
    """이미지 영역에서 표 추출"""
    try:
        # 이미지 전처리
        binary = preprocess_image(image)
        
        # 셀 감지
        cells = detect_table_cells(binary)
        if not cells:
            return {'success': False, 'error': '표를 감지할 수 없습니다'}
        
        # 격자 구조 감지
        rows, cols = detect_grid_structure(cells)
        
        # 셀들을 격자에 할당
        grid = assign_cells_to_grid(cells, rows, cols)
        
        # OCR 수행
        table_data = []
        for row_index, row_cells in enumerate(grid):
            row_data = []
            for col_index, cell_coords in enumerate(row_cells):  # cell -> cell_coords 로 명확화
                if cell_coords:  # 빈 셀이 아닌 경우에만 처리
                    try:
                        # extract_text_from_cell 호출 시 영역 이미지와 셀 좌표를 전달
                        text = extract_text_from_cell(image, cell_coords)  # 수정된 부분
                        row_data.append(text.strip())
                    except Exception as cell_error:
                        logger.error(f"셀 (행 {row_index}, 열 {col_index}) 처리 중 오류: {cell_error}")
                        row_data.append("[오류]")  # 오류 발생 시 표시
                else:
                    row_data.append('')  # 빈 셀은 빈 문자열로 추가
            table_data.append(row_data)
        
        return {
            'success': True,
            'table_data': table_data
        }
        
    except Exception as e:
        # 상세한 오류 로깅
        logger.exception(f"extract_table_from_image_region 함수에서 오류 발생: {str(e)}")
        return {'success': False, 'error': str(e)}


def extract_text_from_region(image: np.ndarray) -> str:
    """이미지 영역에서 텍스트 추출"""
    reader = get_ocr_reader()
    results = reader.readtext(image, detail=0)
    return ' '.join(results) if results else ''


def save_content_to_files(content: List[Dict], base_path: str) -> Dict[str, Any]:
    """추출된 콘텐츠를 파일로 저장 - 개선된 버전"""
    try:
        # CSV 파일 생성
        csv_path = f"{base_path}_content.csv"
        excel_path = f"{base_path}_content.xlsx"
        
        # 데이터 변환
        rows = []
        current_row = []
        
        for element in content:
            if element['type'] == 'text':
                # 텍스트 요소는 새로운 행으로 추가
                if current_row:
                    rows.append(current_row)
                    current_row = []
                rows.append([element['data']])
            else:  # table
                # 표 데이터는 그대로 추가
                if current_row:
                    rows.append(current_row)
                    current_row = []
                rows.extend(element['data'])
        
        if current_row:
            rows.append(current_row)
        
        # CSV 저장
        save_table_to_csv(rows, csv_path)
        
        # Excel 저장
        save_table_to_excel(rows, excel_path)
        
        return {
            'success': True,
            'csv_path': csv_path,
            'excel_path': excel_path
        }
        
    except Exception as e:
        logger.error(f"파일 저장 중 오류 발생: {str(e)}")
        return {'success': False, 'error': str(e)}


def evaluate_table_quality(table_data):
    """추출된 표의 품질 점수를 계산"""
    if not table_data or len(table_data) < 2:
        return 0.0, "데이터가 충분하지 않음"
    
    score = 1.0  # 기본 점수
    issues = []
    
    # 1. 헤더 품질 체크
    header = table_data[0]
    if not header:
        score -= 0.2
        issues.append("헤더 없음")
    else:
        # 1.1 자동 생성된 헤더인지 확인
        auto_headers = sum(1 for h in header if h.startswith("Column_"))
        if auto_headers > 0:
            score -= 0.1 * (auto_headers / len(header))
            issues.append(f"자동 생성된 헤더 {auto_headers}/{len(header)}")
        
        # 1.2 중복 헤더 체크
        unique_headers = len(set(header))
        if unique_headers < len(header):
            score -= 0.05
            issues.append(f"중복 헤더 {len(header) - unique_headers}개")
    
    # 2. 데이터 행의 일관성 체크
    data_rows = table_data[1:]
    if not data_rows:
        score -= 0.3
        issues.append("데이터 행 없음")
    else:
        # 2.1 길이 일관성
        row_lengths = [len(row) for row in data_rows]
        if len(set(row_lengths)) > 1:
            score -= 0.1
            issues.append("일부 행의 길이가 일치하지 않음")
        
        # 2.2 빈 셀 비율
        total_cells = sum(row_lengths)
        empty_cells = sum(sum(1 for cell in row if not cell) for row in data_rows)
        empty_ratio = empty_cells / total_cells if total_cells > 0 else 0
        if empty_ratio > 0.2:
            score -= 0.1 * min(empty_ratio, 0.5) / 0.2
            issues.append(f"빈 셀 비율 높음 ({empty_ratio:.0%})")
        
        # 2.3 연속 중복 행 체크
        duplicate_rows = 0
        for i in range(1, len(data_rows)):
            if data_rows[i] == data_rows[i-1]:
                duplicate_rows += 1
        if duplicate_rows > 0:
            score -= 0.05 * min(duplicate_rows / len(data_rows), 0.5) / 0.25
            issues.append(f"중복 행 {duplicate_rows}개")
    
    # 점수 범위 제한
    score = max(0.0, min(1.0, score))
    return score, ", ".join(issues) if issues else "양호"


def extract_and_save_table(file_path: str, output_format: str = TableConfig.DEFAULT_OUTPUT_FORMAT) -> Dict[str, Any]:
    """파일에서 표 추출 및 저장 (이미지 또는 PDF)"""
    # 파일 확장자 확인
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # PDF인 경우 전용 함수 사용
    if file_ext == '.pdf':
        result = extract_table_from_pdf(file_path)
    else:  # 이미지인 경우 이미지 처리 함수 사용
        result = extract_table_from_image(file_path)
    
    if not result.get('success'):
        return result
    
    # 출력 디렉토리 및 파일명 설정
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    output_dir = os.path.dirname(file_path)
    
    table_data = result.get('table_data', [])
    
    # 표 품질 평가
    quality_score, quality_issues = evaluate_table_quality(table_data)
    logger.info(f"표 품질 평가: 점수={quality_score:.2f}, 이슈={quality_issues}")
    
    # CSV 형식으로 저장
    csv_output_path = os.path.join(output_dir, f"{base_name}_table.csv")
    csv_save_success = save_table_to_csv(table_data, csv_output_path)
    
    # Excel 형식으로 저장
    excel_output_path = os.path.join(output_dir, f"{base_name}_table.xlsx")
    excel_save_success = save_table_to_excel(table_data, excel_output_path)
    
    # 지정된 형식에 따라 반환할 출력 경로 결정
    if output_format.lower() == 'excel':
        output_path = excel_output_path
        save_success = excel_save_success
    else:  # CSV (기본값)
        output_path = csv_output_path
        save_success = csv_save_success
    
    if not save_success:
        return {
            'success': False,
            'error': f"표 데이터를 {output_format} 형식으로 저장하지 못했습니다"
        }
    
    return {
        'success': True,
        'table_data': table_data,
        'output_path': output_path,
        'csv_path': csv_output_path,
        'excel_path': excel_output_path,
        'format': output_format,
        'cells_count': result.get('cells_count', result.get('tables_count', 0)),
        'quality_score': quality_score,
        'quality_issues': quality_issues
    } 